"""
    Easy Statistica app
    Created November 2023
    Copyright (C) Raman Lukashevich

    This program is free software; you can redistribute it and/or
    modify it under the terms of the GNU General Public License
    as published by the Free Software Foundation; either version 3
    of the License, or (at your option) any later version.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.
"""

## IMPORTS
import csv
import os
import sys
from math import sqrt
import numpy as np
import openpyxl
import pandas
import scikit_posthocs
from PySide6 import QtGui, QtWidgets
from PySide6.QtCore import QEvent, QFileInfo
from PySide6.QtGui import QGuiApplication
from PySide6.QtPrintSupport import QPrinter
from PySide6.QtWidgets import QFileDialog, QMessageBox, QLineEdit
from fpdf import FPDF
from scipy import stats
from scipy.stats import (chi2, t, poisson, kstest, shapiro, normaltest, ttest_1samp, f_oneway, ttest_rel, ttest_ind,
                         norm, ks_2samp, pearsonr, kendalltau, spearmanr)
from sklearn.linear_model import LinearRegression, LogisticRegression
from sklearn.model_selection import train_test_split
from statsmodels.stats.weightstats import ztest
from statsmodels import api as sm
from sklearn.metrics import roc_curve, roc_auc_score, precision_recall_curve, f1_score, auc

# IMPORT GUI FILE
from DFModel import DFModel
from ui_interface import *


# MAIN WINDOW CLASS
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.setWindowFlag(Qt.FramelessWindowHint)
        self.setWindowOpacity(1)

        # SizeGrip
        self.gripSize = 20
        self.grip = QtWidgets.QSizeGrip(self)
        self.grip.resize(self.gripSize, self.gripSize)

        # move window
        self.ui.HeaderContainer.mouseMoveEvent = self.move_window

        # control Header Container
        self.ui.MinimizeBtn.clicked.connect(self.control_MinimizeBtn)
        self.ui.NormalBtn.clicked.connect(self.control_NormalBtn)
        self.ui.RestoreBtn.clicked.connect(self.control_RestoreBtn)
        self.ui.CloseBtn.clicked.connect(lambda: self.close())
        self.ui.NormalBtn.hide()

        # Hide Work Panel
        self.ui.LeftMenuSubContainer.hide()
        self.ui.CentralMenuSubContainer.hide()
        self.ui.RightContentMenuContainer.hide()
        self.ui.RightMenuSubContainer.hide()

        # Left Menu Page connect
        self.ui.DataBtn.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(0))
        self.ui.DataBtn_2.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(0))
        self.ui.ReportsBtn.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(1))
        self.ui.ReportsBtn_2.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(1))
        self.ui.SettingsBtn.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(2))
        self.ui.SettingsBtn_2.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(2))
        self.ui.HelpBtn.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(3))
        self.ui.HelpBtn_2.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(3))
        self.ui.InfoBtn.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(4))
        self.ui.InfoBtn_2.clicked.connect(lambda: self.ui.stackedWidget.setCurrentIndex(4))

        # Right Menu Page connect
        self.ui.BaseStatisticsBtn.clicked.connect(lambda: self.ui.stackedWidget_2.setCurrentIndex(0))
        self.ui.StatCriteriaBtn.clicked.connect(lambda: self.ui.stackedWidget_2.setCurrentIndex(1))
        self.ui.CorrRegrBtn.clicked.connect(lambda: self.ui.stackedWidget_2.setCurrentIndex(2))
        self.ui.SampleSizeSelectionBtn.clicked.connect(lambda: self.ui.stackedWidget_2.setCurrentIndex(3))
        self.ui.CoefficientsBtn.clicked.connect(lambda: self.ui.stackedWidget_2.setCurrentIndex(4))

        # Data Page
        self.ui.NumberRowsSb.valueChanged.connect(self.Data_Rows)
        self.ui.NumberColumnSb.valueChanged.connect(self.Data_Columns)
        self.ui.CreateTableBtn.clicked.connect(self.Table_Create)
        # Table Header
        self.ui.tableWidget_2.horizontalHeader().setDefaultSectionSize(100)  # Size of Column
        self.ui.tableWidget_2.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)  # Text Align

        header_labels = []
        for i in range(self.ui.tableWidget_2.columnCount()):
            header_label = "Column {}".format(i + 1)
            header_labels.append(header_label)
        self.ui.tableWidget_2.setHorizontalHeaderLabels(header_labels)
        self.ui.tableWidget_2.horizontalHeader().sectionDoubleClicked.connect(self.change_table_header)

        self.ui.tableWidget_2.installEventFilter(self)
        self.ui.LetsGoBtn.clicked.connect(self.lets_start)
        self.ui.OpenTableBtn.clicked.connect(self.xlsx_open_data)
        self.ui.OpenCsvBtn.clicked.connect(self.open_csv_data)

        # Report Page
        self.ui.SaveExcelBtn.clicked.connect(self.saveXSelection)
        self.ui.SaveCsvBtn.clicked.connect(self.saveCSelection)
        self.ui.ExportPdfBtn.clicked.connect(self.pdf_export)
        self.ui.SaveJpgPushButton.clicked.connect(self.jpg_export)
        self.ui.SaveDfPushButton.clicked.connect(self.df_to_csv)
        self.ui.SaveDfPushButton_2.clicked.connect(self.df_to_excel)

        # Base Statistics Page
        self.ui.BaseCalculateBtn.clicked.connect(self.base_Statistics)
        self.ui.MplWidget.hide()
        self.ui.MplWidget_2.hide()
        self.ui.MplWidget_3.hide()

        # Statistical criteria Page
        self.ui.OneTtestRadioButton.hide()
        self.ui.TwoTRadioButton.hide()
        self.ui.PairTRadioButton.hide()
        self.ui.MultiTRadioButton.hide()
        self.ui.label_34.hide()
        self.ui.ExMeanSpinBox.hide()
        self.ui.StatisticalCalculateBtn.clicked.connect(self.stat_Criteria)
        self.ui.OneTtestRadioButton.toggled.connect(self.combo_box_enable_1)
        self.ui.TwoTRadioButton.toggled.connect(self.combo_box_enable_2)
        self.ui.PairTRadioButton.toggled.connect(self.combo_box_enable_2)
        self.ui.FisherCheckBox.toggled.connect(self.combo_box_enable_2)
        self.ui.UtestCheckBox.toggled.connect(self.combo_box_enable_2)
        self.ui.WilcoxCheckBox.toggled.connect(self.combo_box_enable_2)
        self.ui.KolmCheckBox.toggled.connect(self.combo_box_enable_2)
        self.ui.KruWallCheckBox.toggled.connect(self.combo_box_enable_3)
        self.ui.FriCheckBox.toggled.connect(self.combo_box_enable_3)
        self.ui.MultiTRadioButton.toggled.connect(self.combo_box_enable_3)
        self.ui.AnovaCheckBox.toggled.connect(self.combo_box_enable_3)
        self.ui.ttestCheckBox.toggled.connect(self.criteria_show)

        # Corr&regg Page
        self.ui.CorrSpinBox.setEnabled(False)
        self.ui.CorrSpinBox_2.setEnabled(False)
        self.ui.CorrSpinBox_3.setEnabled(False)
        self.ui.RegrSpinBox.setEnabled(False)
        self.ui.RegrSpinBox_2.setEnabled(False)
        self.ui.RegrSpinBox_3.setEnabled(False)
        self.ui.RegrSpinBox_4.setEnabled(False)
        self.ui.RegrSpinBox_5.setEnabled(False)
        self.ui.RegrSpinBox_6.setEnabled(False)
        self.ui.CorrCalculateBtn.clicked.connect(self.corr_regg)
        self.ui.PearsonCorrCheckBox.toggled.connect(self.combo_box_corr_2)
        self.ui.KendalCorrCheckBox.toggled.connect(self.combo_box_corr_2)
        self.ui.SpearCorrCheckBox.toggled.connect(self.combo_box_corr_2)
        self.ui.LinRegCheckBox.toggled.connect(self.combo_box_corr_3)
        self.ui.RoCheckBox.toggled.connect(self.combo_box_corr_3)

        # Setting Page
        self.ui.SetDataRadioButton_2.setChecked(True)

        # Sample Size Page
        self.ui.SampleCalculateBtn.clicked.connect(self.Sample_Result)

        # Coefficient Page
        self.ui.CoeffCalculateBtn.clicked.connect(self.Coeff_Result)

        # Central Work Page
        self.ui.MainContentSubContainer.setCurrentIndex(0)
        self.ui.BaseCalculateBtn.clicked.connect(lambda: self.ui.MainContentSubContainer.setCurrentIndex(1))
        self.ui.StatisticalCalculateBtn.clicked.connect(lambda: self.ui.MainContentSubContainer.setCurrentIndex(1))
        self.ui.CorrCalculateBtn.clicked.connect(lambda: self.ui.MainContentSubContainer.setCurrentIndex(1))
        # Welcome page && FAQ
        self.ui.FAQButton.clicked.connect(self.welcome_page)
        self.ui.FAQButton_2.clicked.connect(self.FAQ_base)
        self.ui.FAQButton_3.clicked.connect(self.FAQ_criteria)
        self.ui.FAQButton_4.clicked.connect(self.FAQ_correlation)
        self.ui.FAQButton_5.clicked.connect(self.FAQ_sample_size)
        self.ui.FAQButton_6.clicked.connect(self.FAQ_coefficient)
        self.ui.firstWelcomeLabel.setTextFormat(Qt.RichText)
        self.ui.firstWelcomeLabel.setWordWrap(True)
        self.ui.firstWelcomeLabel.setAlignment(Qt.AlignJustify | Qt.AlignVCenter)
        self.ui.firstWelcomeLabel.setText("Welcome to Easy Statistica!")
        file = open("FAQ/welcome page.txt", 'r')
        self.ui.viewWelcomeLabel_1.setWordWrap(True)
        self.ui.viewWelcomeLabel_1.setText(str(file.read()))
        self.ui.viewWelcomeLabel_2.setText(" ")
        file.close()
        file2 = open("FAQ/info.txt", 'r')
        self.ui.infoLabel.setWordWrap(True)
        self.ui.infoLabel.setText(str(file2.read()))
        file2.close()

    # Welcome page
    def welcome_page(self):
        self.ui.WorkingPg.hide()
        self.ui.HelloPg.show()
        self.ui.scrollArea_2.setWidgetResizable(True)
        self.ui.MainContentSubContainer.setCurrentIndex(0)
        self.ui.firstWelcomeLabel.setTextFormat(Qt.RichText)
        self.ui.firstWelcomeLabel.setWordWrap(True)
        self.ui.firstWelcomeLabel.setAlignment(Qt.AlignJustify | Qt.AlignVCenter)
        self.ui.firstWelcomeLabel.setText("Welcome to Easy Statistica!")
        file = open("FAQ/welcome page.txt", 'r')
        self.ui.viewWelcomeLabel_1.setWordWrap(True)
        self.ui.viewWelcomeLabel_1.setText(str(file.read()))
        self.ui.viewWelcomeLabel_1.adjustSize()
        self.ui.viewWelcomeLabel_2.setText(" ")
        file.close()

    def FAQ_base(self):
        self.ui.WorkingPg.hide()
        self.ui.HelloPg.show()
        self.ui.scrollArea_2.setWidgetResizable(True)
        self.ui.MainContentSubContainer.setCurrentIndex(0)
        self.ui.firstWelcomeLabel.setText("[FAQ] Base Statistics:")
        file = open("FAQ/FAQ Basic statistica.txt", 'r')
        self.ui.viewWelcomeLabel_1.setWordWrap(True)
        self.ui.viewWelcomeLabel_1.setText(str(file.read()))
        self.ui.viewWelcomeLabel_1.adjustSize()
        file.close()

    def FAQ_criteria(self):
        self.ui.WorkingPg.hide()
        self.ui.HelloPg.show()
        self.ui.scrollArea_2.setWidgetResizable(True)
        self.ui.MainContentSubContainer.setCurrentIndex(0)
        self.ui.firstWelcomeLabel.setText("[FAQ] Statistical criteria: ")
        file = open("FAQ/FAQ Statistical criteria.txt", 'r')
        self.ui.viewWelcomeLabel_1.setWordWrap(True)
        self.ui.viewWelcomeLabel_1.setText(str(file.read()))
        self.ui.viewWelcomeLabel_1.adjustSize()
        file.close()

    def FAQ_correlation(self):
        self.ui.WorkingPg.hide()
        self.ui.HelloPg.show()
        self.ui.scrollArea_2.setWidgetResizable(True)
        self.ui.MainContentSubContainer.setCurrentIndex(0)
        self.ui.firstWelcomeLabel.setText("[FAQ] Correlation and regression:")
        file = open("FAQ/FAQ Correlation and regression.txt", 'r')
        self.ui.viewWelcomeLabel_1.setWordWrap(True)
        self.ui.viewWelcomeLabel_1.setText(str(file.read()))
        self.ui.viewWelcomeLabel_1.adjustSize()
        file.close()

    def FAQ_sample_size(self):
        self.ui.WorkingPg.hide()
        self.ui.HelloPg.show()
        self.ui.scrollArea_2.setWidgetResizable(True)
        self.ui.MainContentSubContainer.setCurrentIndex(0)
        self.ui.firstWelcomeLabel.setText("[FAQ] Sample size selection:")
        file = open("FAQ/FAQ Sample size selection.txt", 'r')
        self.ui.viewWelcomeLabel_1.setWordWrap(True)
        self.ui.viewWelcomeLabel_1.setText(str(file.read()))
        self.ui.viewWelcomeLabel_1.adjustSize()
        file.close()

    def FAQ_coefficient(self):
        self.ui.WorkingPg.hide()
        self.ui.HelloPg.show()
        self.ui.scrollArea_2.setWidgetResizable(True)
        self.ui.MainContentSubContainer.setCurrentIndex(0)
        self.ui.firstWelcomeLabel.setText("[FAQ] Coefficients:")
        file = open("FAQ/FAQ Coefficients.txt", 'r')
        self.ui.viewWelcomeLabel_1.setWordWrap(True)
        self.ui.viewWelcomeLabel_1.setText(str(file.read()))
        self.ui.viewWelcomeLabel_1.adjustSize()
        file.close()

    ## Base Functions
    # Data Frame create
    def table_to_df(self):
        col_count = self.ui.tableWidget_2.columnCount()
        row_count = self.ui.tableWidget_2.rowCount()
        headers = [self.ui.tableWidget_2.horizontalHeaderItem(i).text()
                   for i in range(col_count)]
        df_list = []
        for row in range(row_count):
            df_list2 = []
            for col in range(col_count):
                table_item = self.ui.tableWidget_2.item(row, col)
                df_list2.append('' if table_item is None
                                else str(table_item.text()))
            df_list.append(df_list2)
        df = pandas.DataFrame(df_list, columns=headers)
        df = (df.replace(regex={',': '.'}))
        if self.ui.SetDataRadioButton.isChecked():
            for i in headers:
                df[i] = pandas.to_numeric(df[i], errors='coerce')
                df[i] = df[i].dropna().astype(float)
            df = df.loc[:, (df != df.iloc[0]).any()]
            return df
        if self.ui.SetDataRadioButton_2.isChecked():
            for i in headers:
                df[i] = pandas.to_numeric(df[i], errors='coerce')
                df[i] = df[i].fillna(0).astype(float)
            df = df.loc[:, (df != df.iloc[0]).any()]
            return df
        if self.ui.SetDataRadioButton_3.isChecked():
            for i in headers:
                df[i] = pandas.to_numeric(df[i], errors='coerce')
            df = df.fillna(df.mean().round(3))
            return df
        if self.ui.SetDataRadioButton_4.isChecked():
            for i in headers:
                df[i] = pandas.to_numeric(df[i], errors='coerce')
                df[i] = df[i].interpolate()
            df = df.loc[:, (df != df.iloc[0]).any()]
            return df
        else:
            for i in headers:
                df[i] = pandas.to_numeric(df[i], errors='coerce')
                df[i] = df[i].fillna(0).astype(float)
            df = df.loc[:, (df != df.iloc[0]).any()]
            return df

    # Function for Base Statistica page
    def base_Statistics(self):
        df = self.table_to_df()
        self.clear_result_page()
        self.ui.WorkingPg.show()
        self.ui.HelloPg.hide()
    # Test for "no data"
        rows = self.ui.tableWidget_2.rowCount()
        if not rows:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No data to analysis.')
            self.ui.ResultViewLabel.setText('No data to analysis')
            return
        for row in range(self.ui.tableWidget_2.rowCount()):
            for col in range(self.ui.tableWidget_2.columnCount()):
                if not self.ui.tableWidget_2.item(row, col):
                    msg = QtWidgets.QMessageBox.information(self, 'Attention',
                                                            'No data to analysis or the table has empty columns or cells.')
                    self.ui.ResultViewLabel.setText('No data to analysis or the table has empty columns or cells. '
                                                    '\nRecommendation: Check for empty columns or empty cells at the '
                                                    'end of your table')
                    return

        dt = df.describe().loc[['mean', 'std', 'min', 'max', '25%', '50%', '75%']].round(self.ui.PNumSpinBox_2.value())
        self.ui.ResultViewLabel.setAlignment(Qt.AlignRight)
        self.ui.ResultViewLabel.setText(dt.to_string())
        self.ui.ResultViewLabel.adjustSize()

        # Pearson test
        if self.ui.PearsonCheckBox.isChecked():
            self.ui.ResultViewLabel_2.setText("\n Pearson test: \n")
            if self.ui.tableWidget_2.rowCount() > 7:
                i = 0
                for column in df:
                    test, propability_value = normaltest(df[column])
                    if propability_value < (1 - self.ui.PLevelSpinBox_2.value()):
                        self.ui.ResultViewLabel_2.setText(self.ui.ResultViewLabel_2.text() +
                                                          "\n" + df.columns.values[i] + ":" +
                                                          "\n" + "p-value < " + str(
                            f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}') +
                                                          "\n" + "Hypothesis of normal distribution rejected.")
                    else:
                        self.ui.ResultViewLabel_2.setText(self.ui.ResultViewLabel_2.text() +
                                                          "\n" + df.columns.values[i] + ":" +
                                                          "\n" + "p-value = " + str(
                            f'{propability_value:.{self.ui.PNumSpinBox.value()}f}') +
                                                          "\n" + "Hypothesis of normal distribution accepted.")
                    i += 1
            else:
                msg = QtWidgets.QMessageBox.information(self, 'Attention', 'Not enough data for Pearson test.')
                self.ui.ResultViewLabel_2.setText("\n Pearson test: \n Not enough data for analysis \n")
        self.ui.ResultViewLabel_2.adjustSize()

        # Shapiro test
        if self.ui.ShapiroCheckBox.isChecked():
            self.ui.ResultViewLabel_3.setText("\n Shapiro test: \n")
            i = 0
            for column in df:
                test, propability_value = shapiro(df[column])
                if propability_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    self.ui.ResultViewLabel_3.setText(self.ui.ResultViewLabel_3.text() +
                                                      "\n" + df.columns.values[i] +
                                                      "\n" + "p-value < " + str(
                        f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}') +
                                                      "\n" + "Hypothesis of normal distribution rejected")
                else:
                    self.ui.ResultViewLabel_3.setText(self.ui.ResultViewLabel_3.text() +
                                                      "\n" + df.columns.values[i] +
                                                      "\n" + "p-value = " + str(
                        f'{propability_value:.{self.ui.PNumSpinBox.value()}f}') +
                                                      "\n" + "Hypothesis of normal distribution accepted")
                i += 1

        self.ui.ResultViewLabel_3.adjustSize()

        # Kolmogorov test
        if self.ui.KolmogorovCheckBox.isChecked():
            self.ui.ResultViewLabel_4.setText("\n Kolmogorov test: \n")
            i = 0
            for column in df:
                test, propability_value = kstest(df[column], 'norm')
                if propability_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    self.ui.ResultViewLabel_4.setText(self.ui.ResultViewLabel_4.text() +
                                                      "\n" + df.columns.values[i] +
                                                      "\n" + "p-value < " + str(
                        f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}') +
                                                      "\n" + "Hypothesis of normal distribution rejected")
                else:
                    self.ui.ResultViewLabel_4.setText(self.ui.ResultViewLabel_4.text() +
                                                      "\n" + df.columns.values[i] +
                                                      "\n" + "p-value = " + str(
                        f'{propability_value:.{self.ui.PNumSpinBox.value()}f}') +
                                                      "\n" + "Hypothesis of normal distribution accepted")
                i += 1
        self.ui.ResultViewLabel_4.adjustSize()

        # z-test
        if self.ui.ZCheckBox.isChecked():
            self.ui.ResultViewLabel_5.setText("\n Z test: \n")
            i = 0
            for column in df:
                test, propability_value = ztest(df[column])
                if propability_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    self.ui.ResultViewLabel_5.setText(self.ui.ResultViewLabel_5.text() +
                                                      "\n" + df.columns.values[i] +
                                                      "\n" + "p-value < " + str(
                        f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}') +
                                                      "\n" + "Hypothesis of normal distribution rejected")
                else:
                    self.ui.ResultViewLabel_5.setText(self.ui.ResultViewLabel_5.text() +
                                                      "\n" + df.columns.values[i] +
                                                      "\n" + "p-value = " + str(
                        f'{propability_value:.{self.ui.PNumSpinBox.value()}f}') +
                                                      "\n" + "Hypothesis of normal distribution accepted")
                i += 1

        if (self.ui.ZCheckBox.isChecked() or self.ui.ShapiroCheckBox.isChecked()
                or self.ui.KolmogorovCheckBox.isChecked() or self.ui.PearsonCheckBox.isChecked()):
            self.ui.ResultViewLabel_1.setText('\n Normality test: ')
            self.ui.ResultViewLabel_6.setText(
                '\n If the hypothesis of normality is rejected, '
                '\n it is better to use the median (50% value) instead of the mean value.')
        else:
            self.ui.ResultViewLabel_1.setText('')
            self.ui.ResultViewLabel_2.setText('')
            self.ui.ResultViewLabel_3.setText('')
            self.ui.ResultViewLabel_4.setText('')
            self.ui.ResultViewLabel_5.setText('')
            self.ui.ResultViewLabel_6.setText('')

        # DF Table create
        model = DFModel(df)
        self.ui.dfTableView.show()
        self.ui.dfTableView.setModel(model)
        self.ui.dfTableView.adjustSize()
        self.ui.ResultViewLabel_8.setText('\n Table View: ')

        # Plot create
        if self.ui.BaseCheckBox.isChecked() and self.ui.BaseSpinBox.currentText() == 'All':
            self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
            self.ui.MplWidget.show()
            self.ui.MplWidget.canvas.axes.clear()
            self.ui.MplWidget.canvas.axes.hist(df, edgecolor='k', alpha=0.5)
            self.ui.MplWidget.canvas.axes.legend(df.columns.values, loc='upper right')
            self.ui.MplWidget.canvas.draw()
        if self.ui.BaseCheckBox_2.isChecked() and self.ui.BaseSpinBox_2.currentText() == 'All':
            self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
            self.ui.MplWidget_2.show()
            self.ui.MplWidget_2.canvas.axes.clear()
            self.ui.MplWidget_2.canvas.axes.boxplot(df, labels=df.columns.values)
            self.ui.MplWidget_2.canvas.draw()
        if self.ui.BaseCheckBox.isChecked() and self.ui.BaseSpinBox.currentText() != 'All':
            self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
            anhist = self.ui.BaseSpinBox.currentText()
            anhistlist = df.loc[:, anhist]
            self.ui.MplWidget.show()
            self.ui.MplWidget.canvas.axes.clear()
            self.ui.MplWidget.canvas.axes.hist(anhistlist, density=True, edgecolor='k', alpha=0.5)
            labels = [anhist]
            self.ui.MplWidget.canvas.axes.legend(labels, loc='upper right')
            mu, std = norm.fit(anhistlist)
            xx = np.linspace(min(anhistlist), max(anhistlist), 100)
            p = norm.pdf(xx, mu, std)
            self.ui.MplWidget.canvas.axes.plot(xx, p, 'k', linewidth=2)
            self.ui.MplWidget.canvas.draw()
        if self.ui.BaseCheckBox_2.isChecked() and self.ui.BaseSpinBox_2.currentText() != 'All':
            self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
            anbox = self.ui.BaseSpinBox_2.currentText()
            anboxlist = df.loc[:, anbox]
            self.ui.MplWidget_2.show()
            self.ui.MplWidget_2.canvas.axes.clear()
            self.ui.MplWidget_2.canvas.axes.boxplot(anboxlist)
            self.ui.MplWidget_2.canvas.axes.set_xlabel(anbox)
            self.ui.MplWidget_2.canvas.draw()

    ## Statistical criteria page
    def stat_Criteria(self):
        global group1list, group2list, group3list, group4list, group5list, group1, group2, group3, group4, group5, stat
        rows = self.ui.tableWidget_2.rowCount()
        if not rows:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No data to analysis.')
            self.ui.ResultViewLabel.setText('No data to analysis')
            return
        for row in range(self.ui.tableWidget_2.rowCount()):
            for col in range(self.ui.tableWidget_2.columnCount()):
                if not self.ui.tableWidget_2.item(row, col):
                    msg = QtWidgets.QMessageBox.information(self, 'Attention',
                                                            'No data to analysis or the table has empty columns or cells.')
                    self.ui.ResultViewLabel.setText('No data to analysis or the table has empty columns or cells. '
                                                    '\nRecommendation: Check for empty columns or empty cells at the '
                                                    'end of your table')
                    return
        self.clear_result_page()
        df = self.table_to_df()
        self.ui.WorkingPg.show()
        self.ui.HelloPg.hide()

        # DF Table create
        model = DFModel(df)
        self.ui.dfTableView.show()
        self.ui.dfTableView.setModel(model)
        self.ui.dfTableView.adjustSize()
        self.ui.ResultViewLabel_8.setText('\n Table View: ')
        # Create the spinbox texts
        if self.ui.StatSpinBox.currentText() != ' ':
            group1 = self.ui.StatSpinBox.currentText()
            group1list = df.loc[:, group1]
        if self.ui.StatSpinBox_2.currentText() != ' ':
            group2 = self.ui.StatSpinBox_2.currentText()
            group2list = df.loc[:, group2]
        if self.ui.StatSpinBox_3.currentText() != ' ':
            group3 = self.ui.StatSpinBox_3.currentText()
            group3list = df.loc[:, group3]
        if self.ui.StatSpinBox_4.currentText() != ' ':
            group4 = self.ui.StatSpinBox_4.currentText()
            group4list = df.loc[:, group4]
        if self.ui.StatSpinBox_5.currentText() != ' ':
            group5 = self.ui.StatSpinBox_5.currentText()
            group5list = df.loc[:, group5]
        ExMean = pandas.to_numeric(str(self.ui.ExMeanSpinBox.value()))

        # T-test (one)
        if self.ui.ttestCheckBox.isChecked() and self.ui.OneTtestRadioButton.isChecked():
            self.ui.ResultViewLabel_2.setText("\n Result of One Sample T-Test: ")
            t_statistic, p_value = ttest_1samp(group1list, ExMean)
            if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                self.ui.ResultViewLabel_2.setText(self.ui.ResultViewLabel_2.text() + "\n" + group1 + "\n t_statistic = "
                                                  + str(
                    f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                                                  + str(
                    f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                                                  + '\n Reject Null Hypothesis. \n Expected mean of a sample of '
                                                    ' \n independent observations is not equal \n '
                                                    'to the specified population mean.')
            else:
                self.ui.ResultViewLabel_2.setText(self.ui.ResultViewLabel_2.text() + "\n" + group1 + "\n t_statistic = "
                                                  + str(
                    f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                    f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. \n Expected '
                                                                    'mean of a sample of independent'
                                                                    '\n observations is equal to the specified '
                                                                    'population mean.')
        self.ui.ResultViewLabel.adjustSize()

        # T-test (two)
        if self.ui.ttestCheckBox.isChecked() and self.ui.TwoTRadioButton.isChecked():
            self.ui.ResultViewLabel_2.setText("\n Result of Two Sample T-Test: ")
            t_statistic, p_value = ttest_ind(group1list, group2list)
            if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                    + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                    + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                    + '\n Reject Null Hypothesis. \n Mean of ' + group1 + ' is different from '
                    + group2 + '.')
            else:
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                    + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                        f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. \n Mean of '
                    + group1 + ' is equal to ' + group2 + '.')
            self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
            self.mplwidget_hist()
        # Create the boxplot
            self.ui.MplWidget_2.show()
            labels = [group1, group2]
            self.ui.MplWidget_2.canvas.axes.clear()
            self.ui.MplWidget_2.canvas.axes.boxplot([group1list, group2list])
            self.ui.MplWidget_2.canvas.axes.set_xticklabels(labels)
            self.ui.MplWidget_2.canvas.draw()
        self.ui.ResultViewLabel.adjustSize()

        # T-test (paired)
        if self.ui.ttestCheckBox.isChecked() and self.ui.PairTRadioButton.isChecked():
            self.ui.ResultViewLabel_2.setText("\n Result of Paired T-Test: ")
            t_statistic, p_value = ttest_rel(group1list, group2list)
            if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                    + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                    + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                    + '\n Reject Null Hypothesis. \n Mean of ' + group1 + ' is different from ' + group2 + '.')
            else:
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                    + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                        f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. \n Mean of '
                    + group1 + ' is equal to ' + group2 + '.')

            self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
            self.mplwidget_hist()
        # Create the boxplot
            self.ui.MplWidget_2.show()
            self.ui.MplWidget_2.canvas.axes.clear()
            labels = [group1, group2]
            self.ui.MplWidget_2.canvas.axes.boxplot([group1list, group2list])
            self.ui.MplWidget_2.canvas.axes.set_xticklabels(labels)
            self.ui.MplWidget_2.canvas.draw()
        self.ui.ResultViewLabel.adjustSize()

        # Multi t-test
        if self.ui.ttestCheckBox.isChecked() and self.ui.MultiTRadioButton.isChecked():
            self.ui.ResultViewLabel_2.setText("\n Result of Multiple T-Test : ")
            for a in df.columns:
                for b in df.columns:
                    stat, p = ttest_ind(df[a], df[b])
                    self.ui.ResultViewLabel_2.setText(
                        self.ui.ResultViewLabel_2.text() + "\n " + a + '  ' + b + '  ' + f'{stat:.{self.ui.PNumSpinBox_2.value()}f}, '
                                                                                         f'{p:.{self.ui.PNumSpinBox.value()}f}')
            self.ui.ResultViewLabel_6.setText(
                "\n  Note: \n if p > 0.05: probably the same distribution \n if p < 0.05: probably different distributions \n")

        # ANOVA
        # Check spinbox (2) and calculate statistics
        if self.ui.AnovaCheckBox.isChecked():
            self.ui.ResultViewLabel_3.setText("\n Result of ANOVA: ")
            if (self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' '
                    and self.ui.StatSpinBox_3.currentText() == ' ' and self.ui.StatSpinBox_4.currentText() == ' '
                    and self.ui.StatSpinBox_5.currentText() == ' '):
                t_statistic, p_value = f_oneway(group1list, group2list)
                if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    self.ui.ResultViewLabel_3.setText(
                        self.ui.ResultViewLabel_3.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                        + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                        + '\n Reject Null Hypothesis. \n There is a difference between ' + group1 + ' and ' + group2 + '.')
                else:
                    self.ui.ResultViewLabel_3.setText(
                        self.ui.ResultViewLabel_3.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                            f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                            '\n There is no difference between ' + group1 + ' and ' + group2 + '.')
                self.ui.ResultViewLabel_3.adjustSize()
                self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
                self.mplwidget_hist()
            # Create the boxplot
                self.ui.MplWidget_2.show()
                labels = [group1, group2]
                self.ui.MplWidget_2.canvas.axes.clear()
                self.ui.MplWidget_2.canvas.axes.boxplot([group1list, group2list])
                self.ui.MplWidget_2.canvas.axes.set_xticklabels(labels)
                self.ui.MplWidget_2.canvas.draw()
            # Check spinbox (3) and calculate statistics
            if (self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' '
                    and self.ui.StatSpinBox_3.currentText() != ' ' and self.ui.StatSpinBox_4.currentText() == ' '
                    and self.ui.StatSpinBox_5.currentText() == ' '):
                t_statistic, p_value = f_oneway(group1list, group2list, group3list)
                if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    self.ui.ResultViewLabel_3.setText(
                        self.ui.ResultViewLabel_3.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                        + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                        + '\n Reject Null Hypothesis. \n There is a difference between variables 1 - 3.')
                else:
                    self.ui.ResultViewLabel_3.setText(
                        self.ui.ResultViewLabel_3.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                            f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                            '\n There is no difference between variables 1 - 3.')
                self.ui.ResultViewLabel_3.adjustSize()
                self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
                self.mplwidget_hist()
            # Create the boxplot
                self.ui.MplWidget_2.show()
                labels = [group1, group2, group3]
                self.ui.MplWidget_2.canvas.axes.clear()
                self.ui.MplWidget_2.canvas.axes.boxplot([group1list, group2list, group3list])
                self.ui.MplWidget_2.canvas.axes.set_xticklabels(labels)
                self.ui.MplWidget_2.canvas.draw()
            # Check spinbox (4) and calculate statistics
            if (self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' '
                    and self.ui.StatSpinBox_3.currentText() != ' ' and self.ui.StatSpinBox_4.currentText() != ' '
                    and self.ui.StatSpinBox_5.currentText() == ' '):
                t_statistic, p_value = f_oneway(group1list, group2list, group3list, group4list)
                if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    self.ui.ResultViewLabel_3.setText(
                        self.ui.ResultViewLabel_3.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs '
                        + group4 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                        + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                        + '\n Reject Null Hypothesis. \n There is a difference between variables 1 - 4.')
                else:
                    self.ui.ResultViewLabel_3.setText(
                        self.ui.ResultViewLabel_3.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs '
                        + group4 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                            f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                            '\n There is no difference between variables 1 - 4.')
                self.ui.ResultViewLabel_3.adjustSize()
                self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
                self.mplwidget_hist()
            # Create the boxplot
                self.ui.MplWidget_2.show()
                labels = [group1, group2, group3, group4]
                self.ui.MplWidget_2.canvas.axes.clear()
                self.ui.MplWidget_2.canvas.axes.boxplot([group1list, group2list, group3list, group4list])
                self.ui.MplWidget_2.canvas.axes.set_xticklabels(labels)
                self.ui.MplWidget_2.canvas.draw()
            # Check spinbox (5) and calculate statistics
            if (self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' '
                    and self.ui.StatSpinBox_3.currentText() != ' ' and self.ui.StatSpinBox_4.currentText() != ' '
                    and self.ui.StatSpinBox_5.currentText() != ' '):
                t_statistic, p_value = f_oneway(group1list, group2list, group3list, group4list, group5list)
                if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    self.ui.ResultViewLabel_3.setText(
                        self.ui.ResultViewLabel_3.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs ' + group4 + ' vs ' + group5 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                        + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                        + '\n Reject Null Hypothesis. \n There is a difference between variables 1 - 5.')
                else:
                    self.ui.ResultViewLabel_3.setText(
                        self.ui.ResultViewLabel_3.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs ' + group4 + ' vs ' + group5 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                            f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                            '\n There is no difference between variables 1 - 5.')
                self.ui.ResultViewLabel_3.adjustSize()
                self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
                self.mplwidget_hist()
            # Create the boxplot
                self.ui.MplWidget_2.show()
                labels = [group1, group2, group3, group4, group5]
                self.ui.MplWidget_2.canvas.axes.clear()
                self.ui.MplWidget_2.canvas.axes.boxplot([group1list, group2list, group3list, group4list, group5list])
                self.ui.MplWidget_2.canvas.axes.set_xticklabels(labels)
                self.ui.MplWidget_2.canvas.draw()

        # F-test
        if self.ui.FisherCheckBox.isChecked():
            self.ui.ResultViewLabel_2.setText("\n Result of Fisher's F-Test : ")
            x = np.array(group1list)
            y = np.array(group2list)
            f = np.var(group1list, ddof=1) / np.var(group2list, ddof=1)
            p_value = stats.f.cdf(f, x.size - 1, y.size - 1)
            if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n f = "
                    + str(f'{f:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                    + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                    + '\n Reject Null Hypothesis. \n The variances of ' + group1 + ' and ' + group2 + ' are different.')
            else:
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n f = "
                    + str(f'{f:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                        f'{p_value:.{self.ui.PNumSpinBox.value()}f}')
                    + '\n Fail to Reject Null Hypothesis. \n The variances of ' + group1 + ' and ' + group2 + ' are not different.')
            self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
            self.plot_widget_stat()
        self.ui.ResultViewLabel_2.adjustSize()

        # Nonparametric criteria
        # Mann-Whitney U-test
        if self.ui.UtestCheckBox.isChecked():
            self.ui.ResultViewLabel_4.setText("\n Result of Mann-Whitney U-test: ")
            t_statistic, p_value = stats.mannwhitneyu(group1list, group2list, alternative='two-sided')
            if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                self.ui.ResultViewLabel_4.setText(
                    self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                    + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                    + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                    + '\n Reject Null Hypothesis. \n The ' + group1 + ' and ' + group2 + ' have not the equal median.')
            else:
                self.ui.ResultViewLabel_4.setText(
                    self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                    + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                        f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                        '\n The ' + group1 + ' and ' + group2 + ' have the equal median.')
            self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
            self.mplwidget_hist()
        self.ui.ResultViewLabel_4.adjustSize()

        # Wilcoxon test
        if self.ui.WilcoxCheckBox.isChecked():
            self.ui.ResultViewLabel_4.setText("\n Result of Wilcoxon signed-rank test: ")
            t_statistic, p_value = stats.wilcoxon(group1list, group2list, alternative='two-sided')
            if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                self.ui.ResultViewLabel_4.setText(
                    self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                    + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                    + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                    + '\n Reject Null Hypothesis. \n The difference between ' + group1 + ' and ' + group2
                    + ' does not follow a symmetric distribution around zero.')
            else:
                self.ui.ResultViewLabel_4.setText(
                    self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                    + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                        f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                        '\n The difference between ' + group1 + ' and '
                    + group2 + ' follow ' + '\n a symmetric distribution around zero.')
            self.ui.ResultViewLabel_4.adjustSize()
            self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
            self.mplwidget_hist()

        # Kruskal-Wallis test
        if self.ui.KruWallCheckBox.isChecked():
            if self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' ':
                self.ui.ResultViewLabel_4.setText("\n Result of Kruskal-Wallis test: \n")
                self.ui.ResultViewLabel_4.setText(
                    self.ui.ResultViewLabel_4.text() + "For Kruskal-Wallis test needs 3 or more variables")
            # Check spinbox (3) and calculate statistics
            if (self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' '
                    and self.ui.StatSpinBox_3.currentText() != ' ' and self.ui.StatSpinBox_4.currentText() == ' '
                    and self.ui.StatSpinBox_5.currentText() == ' '):
                self.ui.ResultViewLabel_4.setText("\n Result of Kruskal-Wallis test: ")
                t_statistic, p_value = stats.kruskal(group1list, group2list, group3list)
                data = [group1list, group2list, group3list]
                if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    danna = scikit_posthocs.posthoc_dunn(data, p_adjust='bonferroni')
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                        + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                        + '\n Reject Null Hypothesis. \n At least one population median of ' + group1 + ', '
                        + group2 + ' and ' + group3 + ' is different from the rest.')
                    self.ui.ResultViewLabel_5.setText(
                        '\n Dunn test: \n' + str(danna.round(self.ui.PNumSpinBox_2.value())))
                else:
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                            f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                            '\n The median of ' + group1 + ', ' + group2 + ' and ' + group3 + ' are equal.')
                self.ui.ResultViewLabel_4.adjustSize()
                self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
                self.mplwidget_hist()
            # Check spinbox (4) and calculate statistics
            if (self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' '
                    and self.ui.StatSpinBox_3.currentText() != ' ' and self.ui.StatSpinBox_4.currentText() != ' '
                    and self.ui.StatSpinBox_5.currentText() == ' '):
                self.ui.ResultViewLabel_4.setText("\n Result of Kruskal-Wallis test: ")
                t_statistic, p_value = stats.kruskal(group1list, group2list, group3list, group4list)
                data = [group1list, group2list, group3list, group4list]
                if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    danna = scikit_posthocs.posthoc_dunn(data, p_adjust='bonferroni')
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs '
                        + group4 + ":" + "\n t_statistic = " + str(
                            f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                        + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                        + '\n Reject Null Hypothesis. \n At least one population median of ' + group1 + ' - '
                        + group4 + ' is different from the rest.')
                    self.ui.ResultViewLabel_5.setText(
                        '\n Dunn test: \n' + str(danna.round(self.ui.PNumSpinBox_2.value())))
                else:
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs '
                        + group4 + ":" + "\n t_statistic = " + str(
                            f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                            f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                            '\n The median of ' + group1 + ' - ' + group4 + ' are equal.')
                self.ui.ResultViewLabel_4.adjustSize()
                self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
                self.mplwidget_hist()
            # Check spinbox (5) and calculate statistics
            if (self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' '
                    and self.ui.StatSpinBox_3.currentText() != ' ' and self.ui.StatSpinBox_4.currentText() != ' '
                    and self.ui.StatSpinBox_5.currentText() != ' '):
                self.ui.ResultViewLabel_4.setText("\n Result of Kruskal-Wallis test: ")
                t_statistic, p_value = stats.kruskal(group1list, group2list, group3list, group4list, group5list)
                data = [group1list, group2list, group3list, group4list, group5list]
                if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    danna = scikit_posthocs.posthoc_dunn(data, p_adjust='bonferroni')
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs '
                        + group4 + ' vs ' + group5 + ":" + "\n t_statistic = " + str(
                            f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}')
                        + "\n p_value < " + str(
                            f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                        + '\n Reject Null Hypothesis. \n At least one population median of ' + group1 + ' - '
                        + group5 + ' is different from the rest.')
                    self.ui.ResultViewLabel_5.setText(
                        '\n Dunn test: \n' + str(danna.round(self.ui.PNumSpinBox_2.value())))
                else:
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs '
                        + group4 + ' vs ' + group5 + ":" + "\n t_statistic = " + str(
                            f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}')
                        + "\n p_value = " + str(
                            f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                            '\n The median of ' + group1 + ' - ' + group5 + ' are equal.')
                self.ui.ResultViewLabel_4.adjustSize()
                self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
                self.mplwidget_hist()

        # Friedman test
        if self.ui.FriCheckBox.isChecked():
            if self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' ':
                self.ui.ResultViewLabel_4.setText("\n Result of Friedman test: \n")
                self.ui.ResultViewLabel_4.setText(
                    self.ui.ResultViewLabel_4.text() + "For Friedman test needs 3 or more variables")
            # Check spinbox (3) and calculate statistics
            if (self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' '
                    and self.ui.StatSpinBox_3.currentText() != ' ' and self.ui.StatSpinBox_4.currentText() == ' '
                    and self.ui.StatSpinBox_5.currentText() == ' '):
                self.ui.ResultViewLabel_4.setText("\n Result of Friedman test: ")
                t_statistic, p_value = stats.friedmanchisquare(group1list, group2list, group3list)
                if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                        + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                        + '\n Reject Null Hypothesis. \n At least one population mean of ' + group1 + ' - ' + group3
                        + ' \n is different from the rest.')
                else:
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                            f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                            '\n The mean for ' + group1 + ' - ' + group3 + ' is equal.')
                self.ui.ResultViewLabel_4.adjustSize()
                self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
                self.mplwidget_hist()
            # Check spinbox (4) and calculate statistics
            if (self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' '
                    and self.ui.StatSpinBox_3.currentText() != ' ' and self.ui.StatSpinBox_4.currentText() != ' '
                    and self.ui.StatSpinBox_5.currentText() == ' '):
                self.ui.ResultViewLabel_4.setText("\n Result of Friedman test: ")
                t_statistic, p_value = stats.friedmanchisquare(group1list, group2list, group3list, group4list)
                if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs ' + group4 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                        + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                        + '\n Reject Null Hypothesis. \n At least one population mean of ' + group1 + ' - ' + group4
                        + ' \n is different from the rest.')
                else:
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs ' + group4 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                            f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                            '\n The mean for ' + group1 + ' - ' + group4 + ' is equal.')
                self.ui.ResultViewLabel_4.adjustSize()
                self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
                self.mplwidget_hist()
            # Check spinbox (5) and calculate statistics
            if (self.ui.StatSpinBox.currentText() != ' ' and self.ui.StatSpinBox_2.currentText() != ' '
                    and self.ui.StatSpinBox_3.currentText() != ' ' and self.ui.StatSpinBox_4.currentText() != ' '
                    and self.ui.StatSpinBox_5.currentText() != ' '):
                self.ui.ResultViewLabel_4.setText("\n Result of Friedman test: ")
                t_statistic, p_value = stats.friedmanchisquare(group1list, group2list, group3list, group4list,
                                                               group5list)
                if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs ' + group4 + ' vs ' + group4 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                        + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                        + '\n Reject Null Hypothesis. \n At least one population mean of ' + group1 + ' - ' + group5
                        + ' \n is different from the rest.')
                else:
                    self.ui.ResultViewLabel_4.setText(
                        self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ' vs ' + group3 + ' vs ' + group4 + ' vs ' + group5 + ":" + "\n t_statistic = "
                        + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                            f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                            '\n The mean for ' + group1 + ' - ' + group5 + ' is equal.')
                self.ui.ResultViewLabel_4.adjustSize()
                self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
                self.mplwidget_hist()

        # Kolmogorov-Smirnov test
        if self.ui.KolmCheckBox.isChecked():
            self.ui.ResultViewLabel_4.setText("\n Result of Kolmogorov-Smirnov test: ")
            t_statistic, p_value = ks_2samp(group1list, group2list, alternative='two-sided')
            if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                self.ui.ResultViewLabel_4.setText(
                    self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                    + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                    + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}')
                    + '\n Reject Null Hypothesis. \n The values of ' + group1 + ' and '
                    + group2 + ' do not belong to the same distribution law.')
            else:
                self.ui.ResultViewLabel_4.setText(
                    self.ui.ResultViewLabel_4.text() + "\n " + group1 + ' vs ' + group2 + ":" + "\n t_statistic = "
                    + str(f'{t_statistic:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                        f'{p_value:.{self.ui.PNumSpinBox.value()}f}') + '\n Fail to Reject Null Hypothesis. '
                                                                        '\n The values of ' + group1 + ' and '
                    + group2 + '  \n belong to the same distribution law.')
            self.ui.ResultViewLabel_4.adjustSize()
            self.ui.ResultViewLabel_7.setText("\n Graphics: \n")
            self.mplwidget_hist()

    # Creating MplWidget histogram
    def mplwidget_hist(self):
        self.ui.MplWidget.show()
        self.ui.MplWidget.canvas.axes.clear()
        if self.ui.StatSpinBox.currentText() != ' ':
            self.ui.MplWidget.canvas.axes.hist(group1list, color='c', edgecolor='k', alpha=0.5)
            labels = [group1]
            self.ui.MplWidget.canvas.axes.legend(labels, loc='upper right')
        if self.ui.StatSpinBox_2.currentText() != ' ':
            self.ui.MplWidget.canvas.axes.hist(group2list, color='g', edgecolor='k', alpha=0.5)
            labels = [group1, group2]
            self.ui.MplWidget.canvas.axes.legend(labels, loc='upper right')
        if self.ui.StatSpinBox_3.currentText() != ' ':
            self.ui.MplWidget.canvas.axes.hist(group3list, color='y', edgecolor='k', alpha=0.5)
            labels = [group1, group2, group3]
            self.ui.MplWidget.canvas.axes.legend(labels, loc='upper right')
        if self.ui.StatSpinBox_4.currentText() != ' ':
            self.ui.MplWidget.canvas.axes.hist(group4list, color='m', edgecolor='k', alpha=0.5)
            labels = [group1, group2, group3, group4]
            self.ui.MplWidget.canvas.axes.legend(labels, loc='upper right')
        if self.ui.StatSpinBox_5.currentText() != ' ':
            self.ui.MplWidget.canvas.axes.hist(group5list, color='b', edgecolor='k', alpha=0.5)
            labels = [group1, group2, group3, group4, group5]
            self.ui.MplWidget.canvas.axes.legend(labels, loc='upper right')
            self.ui.MplWidget.canvas.draw()

    # Correlation and regression page
    def corr_regg(self):
        global group1clist, group2clist, group3clist, group4clist, group1c, group2c, group3c, group2plist, group3plist, \
            group2p, group3p, group1plist, group1p, mlr, group4p, group5p, group6p, X, Y, x, y
        self.clear_result_page()
        df = self.table_to_df()
        self.ui.WorkingPg.show()
        self.ui.HelloPg.hide()
        # Create the spinbox texts
        if self.ui.CorrSpinBox_2.currentText() != ' ':
            group2c = self.ui.CorrSpinBox_2.currentText()
            group2clist = df.loc[:, group2c]
        if self.ui.CorrSpinBox_3.currentText() != ' ':
            group3c = self.ui.CorrSpinBox_3.currentText()
            group3clist = df.loc[:, group3c]

        # Kendall corr
        if self.ui.KendalCorrCheckBox.isChecked() and self.ui.CorrSpinBox.currentText() == 'All':
            self.ui.ResultViewLabel.setText("\n Result of Kendall rank correlation: ")
            self.ui.ResultViewLabel.setText(
                self.ui.ResultViewLabel.text() + "\n" + str(df.corr(method='kendall').round(2)))
            self.heatmap()
        self.ui.ResultViewLabel.adjustSize()
        # Check spinbox and calculate statistics
        if self.ui.KendalCorrCheckBox.isChecked() and self.ui.CorrSpinBox_2.currentText() != ' ' and self.ui.CorrSpinBox_3.currentText() != ' ':
            self.ui.ResultViewLabel_2.setText("\n Result of Kendall rank correlation: ")
            kendall_coeff, p_value = kendalltau(group2clist, group3clist)
            if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n" + group2c + ' vs ' + group3c + ' :' + "\n kendall rank coefficient = "
                    + str(f'{kendall_coeff:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                    + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}'))
            else:
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n" + group2c + ' vs ' + group3c + ' :' + "\n kendall rank coefficient = "
                    + str(f'{kendall_coeff:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                        f'{p_value:.{self.ui.PNumSpinBox.value()}f}'))
            self.plot_widget()
        self.ui.ResultViewLabel.adjustSize()

        # Pearson corr
        if self.ui.PearsonCorrCheckBox.isChecked() and self.ui.CorrSpinBox.currentText() == 'All':
            self.ui.ResultViewLabel.setText("\n Result of Pearson correlation: ")
            self.ui.ResultViewLabel.setText(
                self.ui.ResultViewLabel.text() + "\n" + str(
                    df.corr(method='pearson').round(self.ui.PNumSpinBox_2.value())))
            self.heatmap()
        self.ui.ResultViewLabel.adjustSize()
        # Check spinbox and calculate statistics
        if self.ui.PearsonCorrCheckBox.isChecked() and self.ui.CorrSpinBox_2.currentText() != ' ' and self.ui.CorrSpinBox_3.currentText() != ' ':
            self.ui.ResultViewLabel_2.setText("\n Result of Pearson correlation: ")
            pearson_coeff, p_value = pearsonr(group2clist, group3clist)
            if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n" + group2c + ' vs ' + group3c + ' :' + "\n pearson coefficient = "
                    + str(f'{pearson_coeff:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                    + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}'))
            else:
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n" + group2c + ' vs ' + group3c + ' :' + "\n pearson coefficient = "
                    + str(f'{pearson_coeff:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                        f'{p_value:.{self.ui.PNumSpinBox.value()}f}'))
            self.plot_widget()
        self.ui.ResultViewLabel.adjustSize()

        # Spearman corr
        if self.ui.SpearCorrCheckBox.isChecked() and self.ui.CorrSpinBox.currentText() == 'All':
            self.ui.ResultViewLabel.setText("\n Result of Spearman rank correlation: ")
            self.ui.ResultViewLabel.setText(
                self.ui.ResultViewLabel.text() + "\n" + str(
                    df.corr(method='spearman').round(self.ui.PNumSpinBox_2.value())))
            self.heatmap()
        self.ui.ResultViewLabel.adjustSize()
        # Check spinbox and calculate statistics
        if self.ui.SpearCorrCheckBox.isChecked() and self.ui.CorrSpinBox_2.currentText() != ' ' and self.ui.CorrSpinBox_3.currentText() != ' ':
            self.ui.ResultViewLabel_2.setText("\n Result of Spearman rank correlation: ")
            spear_coeff, p_value = spearmanr(group2clist, group3clist)
            if p_value < (1 - self.ui.PLevelSpinBox_2.value()):
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n" + group2c + ' vs ' + group3c + ' :' + "\n spearman rank coefficient = "
                    + str(f'{spear_coeff:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value < "
                    + str(f'{(1 - self.ui.PLevelSpinBox_2.value()):.{self.ui.PNumSpinBox.value()}f}'))
            else:
                self.ui.ResultViewLabel_2.setText(
                    self.ui.ResultViewLabel_2.text() + "\n" + group2c + ' vs ' + group3c + ' :' + "\n spearman rank coefficient = "
                    + str(f'{spear_coeff:.{self.ui.PNumSpinBox_2.value()}f}') + "\n p_value = " + str(
                        f'{p_value:.{self.ui.PNumSpinBox.value()}f}'))
            self.plot_widget()
        self.ui.ResultViewLabel.adjustSize()

        # Linear Regression
        # Create the spinbox texts
        if self.ui.RegrSpinBox.currentText() != ' ':
            group1p = self.ui.RegrSpinBox.currentText()
            group1plist = df.loc[:, group1p]
        if self.ui.RegrSpinBox_2.currentText() != ' ':
            group2p = self.ui.RegrSpinBox_2.currentText()
            group2plist = df.loc[:, group2p]
        if self.ui.RegrSpinBox_3.currentText() != ' ':
            group3p = self.ui.RegrSpinBox_3.currentText()
            group3plist = df.loc[:, group3p]
        if self.ui.RegrSpinBox_4.currentText() != ' ':
            group4p = self.ui.RegrSpinBox_4.currentText()
        if self.ui.RegrSpinBox_5.currentText() != ' ':
            group5p = self.ui.RegrSpinBox_5.currentText()
        if self.ui.RegrSpinBox_6.currentText() != ' ':
            group6p = self.ui.RegrSpinBox_6.currentText()
        # Check spinbox (2) and calculate statistics
        if (self.ui.LinRegCheckBox.isChecked() and self.ui.RegrSpinBox.currentText() != ' '
                and self.ui.RegrSpinBox_2.currentText() != ' '):
            x = np.array(group2plist).reshape(-1, 1)
            y = np.array(group1plist)
            reg = LinearRegression().fit(x, y)
            reg.predict(x)
            self.ui.ResultViewLabel_2.setText("\n Result of linear regression: ")
            self.ui.ResultViewLabel_2.setText(
                self.ui.ResultViewLabel_2.text() + "\n " + ' Intercept: ' + str(
                    reg.intercept_.round(self.ui.PNumSpinBox_2.value())) + "\n Coefficients: " + str(
                    reg.coef_.round(self.ui.PNumSpinBox_2.value())) + '\n')
        # Create the plot
            self.ui.MplWidget.show()
            self.ui.MplWidget.canvas.axes.clear()
            self.ui.MplWidget.canvas.axes.plot(group2plist, group1plist, 'go', alpha=0.5)
            self.ui.MplWidget.canvas.axes.plot(group2plist, reg.predict(x))
            self.ui.MplWidget.canvas.axes.set_xlabel(group2p)
            self.ui.MplWidget.canvas.axes.set_ylabel(group1p)
            self.ui.MplWidget.canvas.draw()
        # Check spinbox (3 - 5) and calculate statistics
        if (self.ui.LinRegCheckBox.isChecked() and self.ui.RegrSpinBox.currentText() != ' '
                and self.ui.RegrSpinBox_2.currentText() != ' ' and self.ui.RegrSpinBox_3.currentText() != ' '):
            data = df[[group2p, group3p]]
            x = data
            y = df[group1p]
            self.lin_regr()
        if (self.ui.LinRegCheckBox.isChecked() and self.ui.RegrSpinBox.currentText() != ' '
                and self.ui.RegrSpinBox_2.currentText() != ' ' and self.ui.RegrSpinBox_3.currentText() != ' '
                and self.ui.RegrSpinBox_4.currentText() != ' '):
            data = df[[group2p, group3p, group4p]]
            x = data
            y = df[group1p]
            self.lin_regr()
        if (self.ui.LinRegCheckBox.isChecked() and self.ui.RegrSpinBox.currentText() != ' '
                and self.ui.RegrSpinBox_2.currentText() != ' ' and self.ui.RegrSpinBox_3.currentText() != ' '
                and self.ui.RegrSpinBox_4.currentText() != ' ' and self.ui.RegrSpinBox_5.currentText() != ' '):
            data = df[[group2p, group3p, group4p, group5p]]
            x = data
            y = df[group1p]
            self.lin_regr()
        if (self.ui.LinRegCheckBox.isChecked() and self.ui.RegrSpinBox.currentText() != ' '
                and self.ui.RegrSpinBox_2.currentText() != ' ' and self.ui.RegrSpinBox_3.currentText() != ' '
                and self.ui.RegrSpinBox_4.currentText() != ' ' and self.ui.RegrSpinBox_5.currentText() != ' '
                and self.ui.RegrSpinBox_6.currentText() != ' '):
            data = df[[group2p, group3p, group4p, group5p, group6p]]
            x = data
            y = df[group1p]
            self.lin_regr()

        # Logistical Regression + ROC-curve
        if (self.ui.RoCheckBox.isChecked() and self.ui.RegrSpinBox.currentText() != ' '
                and self.ui.RegrSpinBox_2.currentText() != ' '):
            X = df[[group2p]]
            Y = df[group1p]
            self.roc_analyze()
        if (self.ui.RoCheckBox.isChecked() and self.ui.RegrSpinBox.currentText() != ' '
                and self.ui.RegrSpinBox_2.currentText() != ' ' and self.ui.RegrSpinBox_3.currentText() != ' '):
            X = df[[group2p, group3p]]
            Y = df[group1p]
            self.roc_analyze()
        if (self.ui.RoCheckBox.isChecked() and self.ui.RegrSpinBox.currentText() != ' '
                and self.ui.RegrSpinBox_2.currentText() != ' ' and self.ui.RegrSpinBox_3.currentText() != ' '
                and self.ui.RegrSpinBox_4.currentText() != ' '):
            X = df[[group2p, group3p, group4p]]
            Y = df[group1p]
            self.roc_analyze()
        if (self.ui.RoCheckBox.isChecked() and self.ui.RegrSpinBox.currentText() != ' '
                and self.ui.RegrSpinBox_2.currentText() != ' ' and self.ui.RegrSpinBox_3.currentText() != ' '
                and self.ui.RegrSpinBox_4.currentText() != ' ' and self.ui.RegrSpinBox_5.currentText() != ' '):
            X = df[[group2p, group3p, group4p, group5p]]
            Y = df[group1p]
            self.roc_analyze()
        if (self.ui.RoCheckBox.isChecked() and self.ui.RegrSpinBox.currentText() != ' '
                and self.ui.RegrSpinBox_2.currentText() != ' ' and self.ui.RegrSpinBox_3.currentText() != ' '
                and self.ui.RegrSpinBox_4.currentText() != ' ' and self.ui.RegrSpinBox_5.currentText() != ' '
                and self.ui.RegrSpinBox_6.currentText() != ' '):
            X = df[[group2p, group3p, group4p, group5p, group6p]]
            Y = df[group1p]
            self.roc_analyze()

    def roc_analyze(self):
        X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.4, random_state=0)
        roc_model = LogisticRegression()
        roc_model.fit(X_train, Y_train)
        ns_probs = [0 for _ in range(len(Y_test))]
        lr_probs = roc_model.predict_proba(X_test)
        lr_probs = lr_probs[:, 1]
        ns_auc = roc_auc_score(Y_test, ns_probs)
        lr_auc = roc_auc_score(Y_test, lr_probs)
        ns_fpr, ns_tpr, _ = roc_curve(Y_test, ns_probs)
        lr_fpr, lr_tpr, _ = roc_curve(Y_test, lr_probs)

        logit_model = sm.Logit(Y_train, sm.add_constant(X_train)).fit()
        # logit_model = sm.Logit(Y_train, X_train).fit()
        r_model = logit_model.summary()

        mlr_diff = pandas.DataFrame(
            {'Actual value': Y_test, 'Predicted value': lr_probs.round(0)})
        model = DFModel(mlr_diff)
        self.ui.dfTableView.show()
        self.ui.dfTableView.setModel(model)
        self.ui.ResultViewLabel_8.setText('\n Table View: ')
        self.ui.ResultViewLabel_2.setAlignment(Qt.AlignLeft)
        self.ui.ResultViewLabel_2.setText(str(r_model) + "\n")
        self.ui.ResultViewLabel_2.adjustSize()
        self.ui.ResultViewLabel_7.setText("\n Graphics: ")
        self.ui.MplWidget.show()
        self.ui.MplWidget.canvas.axes.clear()
        self.ui.MplWidget.canvas.axes.plot(lr_fpr, lr_tpr, marker='.',
                                           label="AUC=" + str(lr_auc.round(self.ui.PNumSpinBox_2.value())))
        self.ui.MplWidget.canvas.axes.plot(ns_fpr, ns_tpr, linestyle='--',
                                           label="threshold=" + str(ns_auc.round(self.ui.PNumSpinBox_2.value())))
        self.ui.MplWidget.canvas.axes.set_ylabel('True Positive Rate')
        self.ui.MplWidget.canvas.axes.set_xlabel('False Positive Rate')
        self.ui.MplWidget.canvas.axes.legend(loc=4)
        self.ui.MplWidget.canvas.draw()

        yhat = roc_model.predict(X_test)
        lr_precision, lr_recall, _ = precision_recall_curve(Y_test, lr_probs)
        lr_f1, lr_auc = f1_score(Y_test, yhat), auc(lr_recall, lr_precision)
        no_skill = len(Y_test[Y_test == 1]) / len(Y_test)

        self.ui.MplWidget_2.show()
        self.ui.MplWidget_2.canvas.axes.clear()
        self.ui.MplWidget_2.canvas.axes.plot(lr_recall, lr_precision, marker='.',
                                             label='AUC=' + str(lr_auc.round(self.ui.PNumSpinBox_2.value())))
        self.ui.MplWidget_2.canvas.axes.plot([0, 1], [no_skill, no_skill], linestyle='--',
                                             label="f1_score=" + str(lr_f1.round(self.ui.PNumSpinBox_2.value())))
        self.ui.MplWidget_2.canvas.axes.set_xlabel('Recall')
        self.ui.MplWidget_2.canvas.axes.set_ylabel('Precision')
        self.ui.MplWidget_2.canvas.axes.legend()
        self.ui.MplWidget_2.canvas.draw()

    def lin_regr(self):
        global x
        x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.2)
        mlr = LinearRegression()
        mlr.fit(x_train, y_train)
        y_pred_mlr = mlr.predict(x_test)
        mlr_diff = pandas.DataFrame(
            {'Actual value': y_test, 'Predicted value': y_pred_mlr.round(self.ui.PNumSpinBox_2.value())})
        model = DFModel(mlr_diff)
        self.ui.dfTableView.show()
        self.ui.dfTableView.setModel(model)
        self.ui.ResultViewLabel_8.setText('\n Table View: ')
        x = sm.add_constant(x)
        model = sm.OLS(y, x).fit()
        # predictions = model.predict(x)
        print_model = model.summary()
        self.ui.ResultViewLabel_2.setText("\n Result of linear regression: ")
        self.ui.ResultViewLabel_2.setAlignment(Qt.AlignLeft)
        self.ui.ResultViewLabel_2.setText(self.ui.ResultViewLabel_2.text() + "\n " + str(print_model) + '\n')
        self.ui.MplWidget.hide()

    # Create the heatmap to correlation data
    def heatmap(self):
        global corr_matrix
        df = self.table_to_df()
        if self.ui.PearsonCorrCheckBox.isChecked():
            corr_matrix = df.corr(method='pearson')
        if self.ui.SpearCorrCheckBox.isChecked():
            corr_matrix = df.corr(method='spearman')
        if self.ui.KendalCorrCheckBox.isChecked():
            corr_matrix = df.corr(method='kendall')
        listHeaders = []
        i = 0
        for column in df:
            listHeaders.append(df.columns.values[i])
            i += 1
        self.ui.ResultViewLabel_7.setText("Graphics:")
        self.ui.MplWidget_3.show()
        self.ui.MplWidget_3.canvas.axes.clear()
        self.ui.MplWidget_3.canvas.axes.imshow(corr_matrix, cmap='Pastel1')  # Pastel1
        for i in range(len(listHeaders)):
            for j in range(len(listHeaders)):
                self.ui.MplWidget_3.canvas.axes.annotate(str(round(corr_matrix.values[i][j], 2)),
                                                         xy=(j, i), ha='center', va='center', color='black')

        self.ui.MplWidget_3.canvas.axes.set_xticks(ticks=range(len(listHeaders)), labels=corr_matrix.columns,
                                                   rotation=30)
        self.ui.MplWidget_3.canvas.axes.set_yticks(ticks=range(len(listHeaders)), labels=corr_matrix.columns)
        self.ui.MplWidget_3.canvas.draw()
        self.ui.MplWidget_3.canvas.adjustSize()

    def plot_widget_stat(self):
        self.ui.MplWidget.show()
        self.ui.MplWidget.canvas.axes.clear()
        self.ui.MplWidget.canvas.axes.plot(group1list, group2list, 'go', alpha=0.5)
        self.ui.MplWidget.canvas.axes.set_xlabel(group1)
        self.ui.MplWidget.canvas.axes.set_ylabel(group2)
        self.ui.MplWidget.canvas.draw()

    def plot_widget(self):
        self.ui.MplWidget.show()
        self.ui.MplWidget.canvas.axes.clear()
        self.ui.MplWidget.canvas.axes.plot(group2clist, group3clist, 'go', alpha=0.5)
        self.ui.MplWidget.canvas.axes.set_xlabel(group2c)
        self.ui.MplWidget.canvas.axes.set_ylabel(group3c)
        self.ui.MplWidget.canvas.draw()

    # Functions for correct and beautiful working pages
    def criteria_show(self):
        if self.ui.ttestCheckBox.isChecked():
            self.ui.OneTtestRadioButton.show()
            self.ui.TwoTRadioButton.show()
            self.ui.PairTRadioButton.show()
            self.ui.MultiTRadioButton.show()
            self.ui.label_34.show()
            self.ui.ExMeanSpinBox.show()
        else:
            self.ui.OneTtestRadioButton.hide()
            self.ui.TwoTRadioButton.hide()
            self.ui.PairTRadioButton.hide()
            self.ui.MultiTRadioButton.hide()
            self.ui.label_34.hide()
            self.ui.ExMeanSpinBox.hide()
            self.ui.OneTtestRadioButton.setChecked(False)
            self.ui.TwoTRadioButton.setChecked(False)
            self.ui.PairTRadioButton.setChecked(False)
            self.ui.MultiTRadioButton.setChecked(False)

    def combo_box_enable_1(self):
        self.ui.StatSpinBox_2.setEnabled(False)
        self.ui.StatSpinBox_3.setEnabled(False)
        self.ui.StatSpinBox_4.setEnabled(False)
        self.ui.StatSpinBox_5.setEnabled(False)

    def combo_box_enable_2(self):
        self.ui.StatSpinBox_2.setEnabled(True)
        self.ui.StatSpinBox_3.setEnabled(False)
        self.ui.StatSpinBox_4.setEnabled(False)
        self.ui.StatSpinBox_5.setEnabled(False)
        if self.ui.UtestCheckBox.isChecked():
            self.ui.WilcoxCheckBox.setChecked(False)
            self.ui.KolmCheckBox.setChecked(False)
            self.ui.KruWallCheckBox.setChecked(False)
            self.ui.FriCheckBox.setChecked(False)
        if self.ui.WilcoxCheckBox.isChecked():
            self.ui.UtestCheckBox.setChecked(False)
            self.ui.KolmCheckBox.setChecked(False)
            self.ui.KruWallCheckBox.setChecked(False)
            self.ui.FriCheckBox.setChecked(False)
        if self.ui.KolmCheckBox.isChecked():
            self.ui.UtestCheckBox.setChecked(False)
            self.ui.WilcoxCheckBox.setChecked(False)
            self.ui.KruWallCheckBox.setChecked(False)
            self.ui.FriCheckBox.setChecked(False)
        if self.ui.KruWallCheckBox.isChecked():
            self.ui.UtestCheckBox.setChecked(False)
            self.ui.WilcoxCheckBox.setChecked(False)
            self.ui.KolmCheckBox.setChecked(False)
            self.ui.FriCheckBox.setChecked(False)
        if self.ui.FriCheckBox.isChecked():
            self.ui.UtestCheckBox.setChecked(False)
            self.ui.WilcoxCheckBox.setChecked(False)
            self.ui.KolmCheckBox.setChecked(False)
            self.ui.KruWallCheckBox.setChecked(False)

    def combo_box_enable_3(self):
        self.ui.StatSpinBox_2.setEnabled(True)
        self.ui.StatSpinBox_3.setEnabled(True)
        self.ui.StatSpinBox_4.setEnabled(True)
        self.ui.StatSpinBox_5.setEnabled(True)
        if self.ui.UtestCheckBox.isChecked():
            self.ui.WilcoxCheckBox.setChecked(False)
            self.ui.KolmCheckBox.setChecked(False)
            self.ui.KruWallCheckBox.setChecked(False)
            self.ui.FriCheckBox.setChecked(False)
        if self.ui.WilcoxCheckBox.isChecked():
            self.ui.UtestCheckBox.setChecked(False)
            self.ui.KolmCheckBox.setChecked(False)
            self.ui.KruWallCheckBox.setChecked(False)
            self.ui.FriCheckBox.setChecked(False)
        if self.ui.KolmCheckBox.isChecked():
            self.ui.UtestCheckBox.setChecked(False)
            self.ui.WilcoxCheckBox.setChecked(False)
            self.ui.KruWallCheckBox.setChecked(False)
            self.ui.FriCheckBox.setChecked(False)
        if self.ui.KruWallCheckBox.isChecked():
            self.ui.UtestCheckBox.setChecked(False)
            self.ui.WilcoxCheckBox.setChecked(False)
            self.ui.KolmCheckBox.setChecked(False)
            self.ui.FriCheckBox.setChecked(False)
        if self.ui.FriCheckBox.isChecked():
            self.ui.UtestCheckBox.setChecked(False)
            self.ui.WilcoxCheckBox.setChecked(False)
            self.ui.KolmCheckBox.setChecked(False)
            self.ui.KruWallCheckBox.setChecked(False)

    def combo_box_corr_2(self):
        self.ui.CorrSpinBox.setEnabled(True)
        self.ui.CorrSpinBox_2.setEnabled(True)
        self.ui.CorrSpinBox_3.setEnabled(True)
        if self.ui.PearsonCorrCheckBox.isChecked():
            self.ui.KendalCorrCheckBox.setChecked(False)
            self.ui.SpearCorrCheckBox.setChecked(False)
        if self.ui.SpearCorrCheckBox.isChecked():
            self.ui.KendalCorrCheckBox.setChecked(False)
            self.ui.PearsonCorrCheckBox.setChecked(False)
        if self.ui.KendalCorrCheckBox.isChecked():
            self.ui.SpearCorrCheckBox.setChecked(False)
            self.ui.PearsonCorrCheckBox.setChecked(False)

    def combo_box_corr_3(self):
        self.ui.RegrSpinBox.setEnabled(True)
        self.ui.RegrSpinBox_2.setEnabled(True)
        self.ui.RegrSpinBox_3.setEnabled(True)
        self.ui.RegrSpinBox_4.setEnabled(True)
        self.ui.RegrSpinBox_5.setEnabled(True)
        self.ui.RegrSpinBox_6.setEnabled(True)

    def clear_result_page(self):
        self.ui.ResultViewLabel_1.setText('')
        self.ui.ResultViewLabel.setText('')
        self.ui.ResultViewLabel_2.setText('')
        self.ui.ResultViewLabel_3.setText('')
        self.ui.ResultViewLabel_4.setText('')
        self.ui.ResultViewLabel_5.setText('')
        self.ui.ResultViewLabel_6.setText('')
        self.ui.ResultViewLabel_7.setText('')
        self.ui.ResultViewLabel_8.setText('')
        self.ui.MplWidget.hide()
        self.ui.MplWidget_2.hide()
        self.ui.MplWidget_3.hide()
        self.ui.dfTableView.hide()

    # bridge function to start working with data from a table
    def lets_start(self):
        rows = self.ui.tableWidget_2.rowCount()
        if not rows:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No data to analysis.')
            self.ui.ResultViewLabel.setText('No data to analysis')
            return

        self.clear_result_page()
        df = self.table_to_df()
        listHeaders = []
        i = 0
        for column in df:
            listHeaders.append(df.columns.values[i])
            i += 1
        if len(listHeaders) > 1:
            listHeaders.insert(0, " ")
            self.ui.StatSpinBox.clear()
            self.ui.StatSpinBox.addItems(list(listHeaders))
            self.ui.StatSpinBox_2.clear()
            self.ui.StatSpinBox_2.addItems(list(listHeaders))
            self.ui.StatSpinBox_3.clear()
            self.ui.StatSpinBox_3.addItems(list(listHeaders))
            self.ui.StatSpinBox_4.clear()
            self.ui.StatSpinBox_4.addItems(list(listHeaders))
            self.ui.StatSpinBox_5.clear()
            self.ui.StatSpinBox_5.addItems(list(listHeaders))
            self.ui.RegrSpinBox.clear()
            self.ui.RegrSpinBox.addItems(list(listHeaders))
            self.ui.RegrSpinBox_2.clear()
            self.ui.RegrSpinBox_2.addItems(list(listHeaders))
            self.ui.RegrSpinBox_3.clear()
            self.ui.RegrSpinBox_3.addItems(list(listHeaders))
            self.ui.RegrSpinBox_4.clear()
            self.ui.RegrSpinBox_4.addItems(list(listHeaders))
            self.ui.RegrSpinBox_5.clear()
            self.ui.RegrSpinBox_5.addItems(list(listHeaders))
            self.ui.RegrSpinBox_6.clear()
            self.ui.RegrSpinBox_6.addItems(list(listHeaders))
        else:
            self.ui.StatSpinBox.clear()
            self.ui.StatSpinBox.addItems(list(listHeaders))

        listHeaders = []
        i = 0
        for column in df:
            listHeaders.append(df.columns.values[i])
            i += 1
        if len(listHeaders) > 1:
            listHeaders.insert(0, " ")
            self.ui.CorrSpinBox_2.clear()
            self.ui.CorrSpinBox_2.addItems(list(listHeaders))
            self.ui.CorrSpinBox_3.clear()
            self.ui.CorrSpinBox_3.addItems(list(listHeaders))
            listHeaders = []
            listHeaders.insert(0, " ")
            listHeaders.insert(1, "All")
            self.ui.CorrSpinBox.clear()
            self.ui.CorrSpinBox.addItems(list(listHeaders))
        else:
            pass

        del listHeaders[:]
        listHeaders = []
        i = 0
        for column in df:
            listHeaders.append(df.columns.values[i])
            i += 1

        if len(listHeaders) > 1:
            listHeaders.insert(0, "All")
            self.ui.BaseSpinBox.clear()
            self.ui.BaseSpinBox.addItems(list(listHeaders))
            self.ui.BaseSpinBox_2.clear()
            self.ui.BaseSpinBox_2.addItems(list(listHeaders))
        else:
            self.ui.BaseSpinBox.clear()
            self.ui.BaseSpinBox.addItems(list(listHeaders))
        del listHeaders[:]
        self.ui.RightMenuSubContainer.showMaximized()
        self.ui.CentralMenuSubContainer.hide()

    # Data Functions
    # Creating the table
    def Data_Rows(self):
        rowTable = self.ui.tableWidget_2.rowCount()
        rowSpinBox = self.ui.NumberRowsSb.value()
        if rowTable + 1 == rowSpinBox:
            self.ui.tableWidget_2.insertRow(rowSpinBox - rowTable + 1)
        if rowTable - 1 == rowSpinBox:
            self.ui.tableWidget_2.removeRow(rowTable - rowSpinBox)
        self.ui.tableWidget_2.horizontalHeader().setDefaultSectionSize(100)

    def Data_Columns(self):
        i = 1
        columnTable = self.ui.tableWidget_2.columnCount()
        columnSpinBox = self.ui.NumberColumnSb.value()
        if columnTable + 1 == columnSpinBox:
            col = self.ui.tableWidget_2.columnCount()
            item = QtWidgets.QTableWidgetItem("Column {}".format(col + 1))
            for i in range(col, col + 1):
                self.ui.tableWidget_2.insertColumn(col)
                self.ui.tableWidget_2.setHorizontalHeaderItem(i, item)

        if columnTable - 1 == columnSpinBox:
            col = self.ui.tableWidget_2.columnCount()
            for i in range(col - 1, col):
                self.ui.tableWidget_2.removeColumn(col - 1)
        self.ui.tableWidget_2.horizontalHeader().setDefaultSectionSize(100)

    def Table_Create(self):
        self.ui.tableWidget_2.clear()
        header_labels = []
        for i in range(self.ui.tableWidget_2.columnCount()):
            header_label = "Column {}".format(i + 1)
            header_labels.append(header_label)
        self.ui.tableWidget_2.setHorizontalHeaderLabels(header_labels)
        i = 0
        while self.ui.tableWidget_2.rowCount() < self.ui.NumberRowsSb.value():
            self.ui.tableWidget_2.insertRow(i)
        while self.ui.tableWidget_2.rowCount() > self.ui.NumberRowsSb.value():
            self.ui.tableWidget_2.removeRow(i)
        while self.ui.tableWidget_2.columnCount() < self.ui.NumberColumnSb.value():
            col = self.ui.tableWidget_2.columnCount()
            item = QtWidgets.QTableWidgetItem("Column {}".format(col + 1))
            for i in range(col, col + 1):
                self.ui.tableWidget_2.insertColumn(col)
                self.ui.tableWidget_2.setHorizontalHeaderItem(i, item)
        while self.ui.tableWidget_2.columnCount() > self.ui.NumberColumnSb.value():
            col = self.ui.tableWidget_2.columnCount()
            for i in range(col - 1, col):
                self.ui.tableWidget_2.removeColumn(col - 1)
        self.ui.tableWidget_2.horizontalHeader().setDefaultSectionSize(100)

    def change_table_header(self, index):
        oldHeader = self.ui.tableWidget_2.horizontalHeaderItem(index).text()
        newHeader, ok = QtWidgets.QInputDialog.getText(self,
                                                       'Change header label',
                                                       'Header:',
                                                       QLineEdit.Normal,
                                                       oldHeader)
        if ok:
            self.ui.tableWidget_2.horizontalHeaderItem(index).setText(newHeader)

    def xlsx_open_data(self):
        path, _filter = QtWidgets.QFileDialog.getOpenFileName(self, "Export Excel", None, "Excel files (*.xlsx)")
        if not path:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No open file specified.')
            return
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        self.ui.tableWidget_2.setRowCount(sheet.max_row - 1)
        self.ui.tableWidget_2.setColumnCount(sheet.max_column)

        list_values = list(sheet.values)
        self.ui.tableWidget_2.setHorizontalHeaderLabels(list_values[0])

        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                self.ui.tableWidget_2.setItem(row_index, col_index, QtWidgets.QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1

        for row in range(self.ui.tableWidget_2.rowCount()):
            for col in range(self.ui.tableWidget_2.columnCount()):
                empty = self.ui.tableWidget_2.item(row, col)
                if empty is None:
                    self.ui.tableWidget_2.removeRow(row)

    def xlsx_save_data(self):
        rows = self.ui.tableWidget_2.rowCount()
        if not rows:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No data to save.')
            return
        for row in range(self.ui.tableWidget_2.rowCount()):
            for col in range(self.ui.tableWidget_2.columnCount()):
                if self.ui.tableWidget_2.item(row, col) is None:
                    msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No data to save or empty lines.')
                    return

        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save XLSX', os.getenv('HOME'),
                                                        "Excel files (*.xlsx)")
        if path != '':
            if QFileInfo(path).suffix() == "": path += '.xlsx'
        if not path:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No save file specified.')
            return
        columnHeaders = []

        for j in range(self.ui.tableWidget_2.model().columnCount()):
            columnHeaders.append(self.ui.tableWidget_2.horizontalHeaderItem(j).text())

        df = pandas.DataFrame(columns=columnHeaders)

        for row in range(self.ui.tableWidget_2.rowCount()):
            for col in range(self.ui.tableWidget_2.columnCount()):
                df.at[row, columnHeaders[col]] = self.ui.tableWidget_2.item(row, col).text()

        df.to_excel(path, index=True)
        msg = QtWidgets.QMessageBox.information(self, 'Ok', 'Excel file exported')

    def open_csv_data(self):
        path, ok = QtWidgets.QFileDialog.getOpenFileName(
            self, 'Open CSV', os.getenv('HOME'), 'CSV(*.csv)')
        if not path:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No open file specified.')
            return
        if ok:
            self.ui.tableWidget_2.clear()
            self.ui.tableWidget_2.setRowCount(0)
            with open(path) as csvfile:
                reader = csv.reader(csvfile)
                header = next(reader)
                self.ui.tableWidget_2.setColumnCount(len(header))
                self.ui.tableWidget_2.setHorizontalHeaderLabels(header)
                for row, values in enumerate(reader):
                    self.ui.tableWidget_2.insertRow(row)
                    for column, value in enumerate(values):
                        self.ui.tableWidget_2.setItem(
                            row, column, QtWidgets.QTableWidgetItem(value))

    def save_csv_data(self):
        for row in range(self.ui.tableWidget_2.rowCount()):
            for col in range(self.ui.tableWidget_2.columnCount()):
                if self.ui.tableWidget_2.item(row, col) is None:
                    msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No data to save or empty lines.')
                    return
        path, ok = QtWidgets.QFileDialog.getSaveFileName(
            self, 'Save CSV', os.getenv('HOME'), 'CSV(*.csv)')
        if not path:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No save file specified.')
            return
        if ok:
            columns = range(self.ui.tableWidget_2.columnCount())
            header = [self.ui.tableWidget_2.horizontalHeaderItem(column).text()
                      for column in columns]
            with open(path, 'w') as csvfile:
                writer = csv.writer(
                    csvfile, dialect='excel', lineterminator='\n')
                writer.writerow(header)
                for row in range(self.ui.tableWidget_2.rowCount()):
                    writer.writerow(
                        self.ui.tableWidget_2.item(row, column).text()
                        for column in columns)
        msg = QtWidgets.QMessageBox.information(self, 'Ok', 'CSV file exported')

    def df_to_csv(self):
        if self.ui.dfTableView.isVisible():
            path, ok = QtWidgets.QFileDialog.getSaveFileName(
                self, 'Save CSV', os.getenv('HOME'), 'CSV(*.csv)')
            if not path:
                msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No save file specified.')
                return
            if ok:
                df = self.table_to_df()
                df.to_csv(path, sep='\t', encoding='utf-8')
                msg = QtWidgets.QMessageBox.information(self, 'Ok', 'CSV file exported')
        else:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No data to save')
            return

    def df_to_excel(self):
        if self.ui.dfTableView.isVisible():
            path, ok = QtWidgets.QFileDialog.getSaveFileName(
                self, 'Save Excel', os.getenv('HOME'), 'Excel(*.xlsx)')
            if path != '':
                if QFileInfo(path).suffix() == "": path += '.xlsx'
            if not path:
                msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No save file specified.')
                return
            if ok:
                df = self.table_to_df()
                df.to_excel(path, index=False)
                msg = QtWidgets.QMessageBox.information(self, 'Ok', 'Excel file exported')
        else:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No data to save')
            return

    # CTR+C - CTR+V table function
    def keyPressEvent(self, event):
        super().keyPressEvent(event)
        if event.key() == Qt.Key.Key_C and (event.modifiers() & Qt.KeyboardModifier.ControlModifier):
            copied_cells = sorted(self.ui.tableWidget_2.selectedIndexes())
            copy_text = ''
            max_column = copied_cells[-1].column()
            for c in copied_cells:
                copy_text += self.ui.tableWidget_2.item(c.row(), c.column()).text()
                if c.column() == max_column:
                    copy_text += '\n'
                else:
                    copy_text += '\t'
            QApplication.clipboard().setText(copy_text)

        if event.key() == Qt.Key.Key_V and (event.modifiers() & Qt.KeyboardModifier.ControlModifier):
            selection = self.ui.tableWidget_2.selectedIndexes()
            rows = QApplication.clipboard().text().split('\n')[:-1]
            if self.ui.tableWidget_2.rowCount() < len(rows):
                self.ui.tableWidget_2.setRowCount(len(rows))
            else:
                pass

            self.ui.tableWidget_2.setColumnCount(len(rows[0].split('\t')))

            if selection:
                row_anchor = selection[0].row()
                column_anchor = selection[0].column()
                clipboard = QApplication.clipboard()
                rows = clipboard.text().split('\n')
                for index_row, row in enumerate(rows):
                    values = row.split('\t')
                    for index_col, value in enumerate(values):
                        item = QTableWidgetItem(value)
                        self.ui.tableWidget_2.setItem(row_anchor + index_row, column_anchor + index_col, item)
            header_labels = []
            for i in range(self.ui.tableWidget_2.columnCount()):
                header_label = "Column {}".format(i + 1)
                header_labels.append(header_label)
            self.ui.tableWidget_2.setHorizontalHeaderLabels(header_labels)
        if event.key() in (Qt.Key_Backspace, Qt.Key_Delete):
            self.clearSelection()
        super().keyPressEvent(event)

    #creating a context menu
    def eventFilter(self, source, event):
        if event.type() == QEvent.ContextMenu and source is self.ui.tableWidget_2:
            if event.type() == QEvent.ContextMenu:
                menu = QtWidgets.QMenu()

                openXAction = menu.addAction("Open Excel file")
                openCAction = menu.addAction("Open CSV file")
                saveXAction = menu.addAction("Save to Excel file")
                saveCAction = menu.addAction("Save to CSV file")
                copyAction = menu.addAction("Copy rows")
                pasteAction = menu.addAction("Paste rows")
                clearAction = menu.addAction("Clear rows/columns")
                addRowsAction = menu.addAction("Add Row")
                addColAction = menu.addAction("Add Column")
                delRowsAction = menu.addAction("Delete Row")
                delColAction = menu.addAction("Delete Column")

                openXAction.triggered.connect(self.openXSelection)
                openCAction.triggered.connect(self.openCSelection)
                saveXAction.triggered.connect(self.saveXSelection)
                saveCAction.triggered.connect(self.saveCSelection)
                copyAction.triggered.connect(self.copySelection)
                pasteAction.triggered.connect(self.pasteSelection)
                clearAction.triggered.connect(self.clearSelection)
                addRowsAction.triggered.connect(self.addRowSelection)
                addColAction.triggered.connect(self.addColSelection)
                delRowsAction.triggered.connect(self.delRowSelection)
                delColAction.triggered.connect(self.delColSelection)

                menu.exec(event.globalPos())
                return True
        return super().eventFilter(source, event)

    def openXSelection(self):
        self.xlsx_open_data()

    def openCSelection(self):
        self.open_csv_data()

    def saveXSelection(self):
        self.xlsx_save_data()

    def saveCSelection(self):
        self.save_csv_data()

    def copySelection(self):
        copied_cells = sorted(self.ui.tableWidget_2.selectedIndexes())
        copy_text = ''
        max_column = copied_cells[-1].column()
        for c in copied_cells:
            copy_text += self.ui.tableWidget_2.item(c.row(), c.column()).text()
            if c.column() == max_column:
                copy_text += '\n'
            else:
                copy_text += '\t'
        QApplication.clipboard().setText(copy_text)

    def pasteSelection(self):
        if not QGuiApplication.clipboard():
            return
        selection = self.ui.tableWidget_2.selectedIndexes()
        rows = QApplication.clipboard().text().split('\n')[:-1]
        if self.ui.tableWidget_2.rowCount() < len(rows):
            self.ui.tableWidget_2.setRowCount(len(rows))

        if selection:
            row_anchor = selection[0].row()
            column_anchor = selection[0].column()
            clipboard = QApplication.clipboard()
            rows = clipboard.text().split('\n')[:-1]
            for index_row, row in enumerate(rows):
                values = row.split('\t')
                for index_col, value in enumerate(values):
                    item = QTableWidgetItem(value)
                    self.ui.tableWidget_2.setItem(row_anchor + index_row, column_anchor + index_col, item)
        header_labels = []
        for i in range(self.ui.tableWidget_2.columnCount()):
            header_label = "Column {}".format(i + 1)
            header_labels.append(header_label)
        self.ui.tableWidget_2.setHorizontalHeaderLabels(header_labels)

    def clearSelection(self):
        if not self.ui.tableWidget_2.selectedIndexes():
            return
        for i_row in range(self.ui.tableWidget_2.selectedRanges()[0].topRow(),
                           self.ui.tableWidget_2.selectedRanges()[0].bottomRow() + 1):
            for i_col in range(self.ui.tableWidget_2.selectedRanges()[0].leftColumn(),
                               self.ui.tableWidget_2.selectedRanges()[0].rightColumn() + 1):
                self.ui.tableWidget_2.setItem(i_row, i_col, QtWidgets.QTableWidgetItem(''))

    def addRowSelection(self):
        row = self.ui.tableWidget_2.rowCount()
        self.ui.tableWidget_2.insertRow(row)

    def addColSelection(self):
        col = self.ui.tableWidget_2.columnCount()
        item = QtWidgets.QTableWidgetItem("Column {}".format(col + 1))
        for i in range(col, col + 1):
            self.ui.tableWidget_2.insertColumn(col)
            self.ui.tableWidget_2.setHorizontalHeaderItem(i, item)

    def delRowSelection(self):
        row = self.ui.tableWidget_2.currentRow()
        if row > -1:
            self.ui.tableWidget_2.removeRow(row)
            self.ui.tableWidget_2.selectionModel().clearCurrentIndex()

    def delColSelection(self):
        col = self.ui.tableWidget_2.currentColumn()
        if col > -1:
            self.ui.tableWidget_2.removeColumn(col)
            self.ui.tableWidget_2.selectionModel().clearCurrentIndex()

    # Print to pdf
    def pdf_export(self):
        fn, _ = QFileDialog.getSaveFileName(self, "Export PDF", None, "PDF files (.pdf); All files()")
        if fn != '':
            if QFileInfo(fn).suffix() == "": fn += '.pdf'
            printer = QPrinter()
            printer.setOutputFileName(fn)
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=10)
            pdf.multi_cell(0, 5, txt=str('Easy Statistica v.0.9.3'), align='R')
            pdf.ln()
            pdf.set_font("Arial", size=14)
            pdf.multi_cell(0, 5, txt=str('Result of data analysis:'), align='C')
            pdf.ln()
            pdf.set_font("Arial", size=12)
            pdf.multi_cell(0, 5, txt=str(self.ui.ResultViewLabel.text()), align='L')
            pdf.ln()
            pdf.multi_cell(0, 5, txt=str(self.ui.ResultViewLabel_2.text()), align='L')
            pdf.ln()
            pdf.multi_cell(0, 5, txt=str(self.ui.ResultViewLabel_3.text()), align='L')
            pdf.ln()
            pdf.multi_cell(0, 5, txt=str(self.ui.ResultViewLabel_4.text()), align='L')
            pdf.ln()
            pdf.multi_cell(0, 5, txt=str(self.ui.ResultViewLabel_5.text()), align='L')
            pdf.ln()
            pdf.multi_cell(0, 5, txt=str(self.ui.ResultViewLabel_6.text()), align='L')
            pdf.ln()
            pdf.output(fn)

    # Print image to jpg
    def jpg_export(self):
        fn, _ = QFileDialog.getSaveFileName(self, "Export Image", None, "Image files (.jpg); All Files()")
        printer = QPrinter()
        printer.setOutputFileName(fn)
        i = 1
        if self.ui.MplWidget.isVisible():
            self.ui.MplWidget.canvas.print_jpg(fn + '_{0:01d}.jpg'.format(i))
            i += 1
        if self.ui.MplWidget_2.isVisible():
            self.ui.MplWidget_2.canvas.print_jpg(fn + '_{0:01d}.jpg'.format(i))
            i += 1
        if self.ui.MplWidget_3.isVisible():
            self.ui.MplWidget_3.canvas.print_jpg(fn + '_{0:01d}.jpg'.format(i))
        if self.ui.MplWidget.isVisible() == False and self.ui.MplWidget_2.isVisible() == False and self.ui.MplWidget_3.isVisible() == False:
            msg = QtWidgets.QMessageBox.information(self, 'Attention', 'No graphics to save.')

    # Sample Size Functions page
    def Sample_Result(self, value):
        z = norm.ppf((1 - self.ui.SampleSpinBox_1.value()) / 2)
        w = norm.ppf((1 - self.ui.SampleSpinBox_4.value()) / 2)
        if self.ui.SampleRadioButton.isChecked():
            value = ((z ** 2) * (self.ui.SampleSpinBox_3.value() / 100) * (
                    (100 - self.ui.SampleSpinBox_3.value()) / 100)
                     / ((self.ui.SampleSpinBox_2.value() / 100) ** 2))
            self.ui.SampleResult_1.setText(str(int(value)))
            self.ui.SampleResult_1.setFont(QtGui.QFont("Times", 10))
            self.ui.SampleResult_1.setAlignment(Qt.AlignCenter)
            self.ui.SampleResult_1.adjustSize()
            self.ui.SampleResult_2.setText(str())
        if self.ui.ErrorRadioButton.isChecked():
            value = abs(w) * sqrt(self.ui.SampleSpinBox_6.value() * (100 - self.ui.SampleSpinBox_6.value())
                                  / self.ui.SampleSpinBox_5.value())
            self.ui.SampleResult_2.setText(str(int(value)))
            self.ui.SampleResult_2.setFont(QtGui.QFont("Times", 10))
            self.ui.SampleResult_2.setAlignment(Qt.AlignCenter)
            self.ui.SampleResult_1.setText(str())

    # Coefficient Functions page
    def Coeff_Result(self, value):
        if self.ui.ChiRadioButton.isChecked():
            value = chi2.ppf(self.ui.PLevelSpinBox.value(), self.ui.NCountCoeffBox.value())
            self.ui.ChiResultLebel.setText(str("%.3f" % value))
            self.ui.ChiResultLebel.setFont(QtGui.QFont("Times", 10))
            self.ui.ChiResultLebel.setAlignment(Qt.AlignCenter)
            self.ui.ChiResultLebel.adjustSize()
            self.ui.TResultLebel.setText(str())
            self.ui.ZResultLebel.setText(str())
            self.ui.PResultLebel.setText(str())

        if self.ui.TRadioButton.isChecked():
            value = t.ppf(self.ui.PLevelSpinBox_3.value(), self.ui.NCountCoeffBox_2.value())
            self.ui.TResultLebel.setText(str("%.3f" % value))
            self.ui.TResultLebel.setFont(QtGui.QFont("Times", 10))
            self.ui.TResultLebel.setAlignment(Qt.AlignCenter)
            self.ui.TResultLebel.adjustSize()
            self.ui.ChiResultLebel.setText(str())
            self.ui.ZResultLebel.setText(str())
            self.ui.PResultLebel.setText(str())

        if self.ui.ZRadioButton.isChecked():
            value = norm.ppf(self.ui.PLevelSpinBox_4.value())
            self.ui.ZResultLebel.setText(str("%.3f" % value))
            self.ui.ZResultLebel.setFont(QtGui.QFont("Times", 10))
            self.ui.ZResultLebel.setAlignment(Qt.AlignCenter)
            self.ui.ZResultLebel.adjustSize()
            self.ui.ChiResultLebel.setText(str())
            self.ui.TResultLebel.setText(str())
            self.ui.PResultLebel.setText(str())

        if self.ui.PRadioButton.isChecked():
            value = 100 * poisson.pmf(self.ui.NCountCoeffBox_3.value(), self.ui.PoissonSpinBox.value())
            self.ui.PResultLebel.setText(str("%.1f" % value))
            self.ui.PResultLebel.setFont(QtGui.QFont("Times", 10))
            self.ui.PResultLebel.setAlignment(Qt.AlignCenter)
            self.ui.PResultLebel.adjustSize()
            self.ui.ChiResultLebel.setText(str())
            self.ui.TResultLebel.setText(str())
            self.ui.ZResultLebel.setText(str())

    # Control Header Container function
    def control_MinimizeBtn(self):
        self.showMinimized()

    def control_NormalBtn(self):
        self.showNormal()
        self.ui.NormalBtn.hide()
        self.ui.RestoreBtn.show()

    def control_RestoreBtn(self):
        self.showMaximized()
        self.ui.RestoreBtn.hide()
        self.ui.NormalBtn.show()

    # SizeGrip
    def resizeEvent(self, event):
        rect = self.rect()
        self.grip.move(rect.right() - self.gripSize, rect.bottom() - self.gripSize)

    def mousePressEvent(self, event):
        self.dragPos = event.globalPosition().toPoint()

    def move_window(self, event):
        if not self.isMaximized():
            if event.buttons() == Qt.LeftButton:
                self.move(self.pos() + event.globalPosition().toPoint() - self.dragPos)
                self.dragPos = event.globalPosition().toPoint()
                event.accept()

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton:
            try:
                currPos = self.mapToGlobal(self.pos())
                newPos = self.mapFromGlobal(currPos)
                self.move(newPos)
            except AttributeError:
                pass
        QWidget.mouseMoveEvent(self, event)

    # Error
    def log_uncaught_exceptions(ex_cls, ex, tb):
        text = '{}: {}:\n'.format(ex_cls.__name__, ex)
        import traceback
        text += ''.join(traceback.format_tb(tb))
        QMessageBox.critical(None, 'Error', text)
        sys.exit(1)

    sys.excepthook = log_uncaught_exceptions


## EXECUTE APP
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    mi_app = MainWindow()
    mi_app.show()
    try:
        sys.exit(app.exec())
    except SystemExit:
        print('Closing Window...')
